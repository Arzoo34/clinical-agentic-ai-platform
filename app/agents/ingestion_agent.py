"""File ingestion agent for clinical trial data."""
from __future__ import annotations

import hashlib
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

from app.agents.base import BaseAgent
from app.core.events import EventType
from app.core.models import AgentResult, FileMetadata, PipelineContext
from app.core.storage import Storage


class IngestionAgent(BaseAgent):
    """Ingests Excel files from ZIP uploads or existing directories."""

    def __init__(self, storage: Storage, **kwargs):
        """Initialize ingestion agent."""
        super().__init__("ingestion_agent", **kwargs)
        self.storage = storage

    def execute(self, context: PipelineContext) -> AgentResult:
        """Execute ingestion."""
        self.log_start(context)
        result = AgentResult(agent_name=self.name)

        try:
            context.ingestion_metadata = self._ingest_files()
            result.data = {"file_count": len(context.ingestion_metadata)}
            self.publish_event(
                EventType.FILE_INGESTED,
                context.session_id,
                data=result.data,
            )
        except Exception as e:
            result.success = False
            result.errors = [str(e)]
            self.log_error(context, str(e))

        self.log_completion(context, result)
        return result

    def _ingest_files(self) -> List[FileMetadata]:
        """Discover and ingest Excel files from raw data directory."""
        metadata_list: List[FileMetadata] = []

        raw_files = self.storage.list_raw_files("**/*.xlsx")
        raw_files.extend(self.storage.list_raw_files("**/*.xls"))

        for file_path in raw_files:
            try:
                metadata = self._process_excel_file(file_path)
                if metadata:
                    metadata_list.append(metadata)
            except Exception as e:
                self.logger.warning(
                    f"Failed to process {file_path}: {e}",
                    context={"file": str(file_path), "error": str(e)},
                )

        return metadata_list

    def _process_excel_file(self, file_path: Path) -> Optional[FileMetadata]:
        """Process a single Excel file and extract metadata."""
        if not file_path.exists():
            return None

        # Extract study and site from path
        study_id = self._extract_study_id(file_path)
        site_id = self._extract_site_id(file_path)

        # Try to load with openpyxl
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            sheet_names = wb.sheetnames

            # Get dimensions from first sheet
            ws = wb.active
            row_count = ws.max_row
            col_count = ws.max_column

            wb.close()
        except Exception as e:
            self.logger.warning(
                f"Could not read Excel structure: {e}",
                context={"file": str(file_path)},
            )
            return None

        # Calculate checksum
        checksum = self._calculate_checksum(file_path)

        # Get file stats
        stat = file_path.stat()
        mod_time = datetime.fromtimestamp(stat.st_mtime).isoformat()

        metadata = FileMetadata(
            filename=file_path.name,
            file_path=str(file_path),
            study_id=study_id,
            site_id=site_id,
            file_type="xlsx" if file_path.suffix.lower() == ".xlsx" else "xls",
            row_count=max(0, row_count - 1),  # Subtract header
            column_count=col_count,
            modification_time=mod_time,
            file_size=stat.st_size,
            checksum=checksum,
            sheet_names=sheet_names,
        )

        return metadata

    def _extract_study_id(self, file_path: Path) -> Optional[str]:
        """Extract study ID from file path."""
        # Look for 'Study X' pattern in path
        for part in file_path.parts:
            if part.lower().startswith("study"):
                return part
        return None

    def _extract_site_id(self, file_path: Path) -> Optional[str]:
        """Extract site ID from file path or name."""
        # Could be in filename or path
        name_parts = file_path.stem.lower().split("_")
        for part in name_parts:
            if part.startswith("site"):
                return part
        return None

    def _calculate_checksum(self, file_path: Path) -> str:
        """Calculate SHA256 checksum of file."""
        sha256_hash = hashlib.sha256()
        with open(file_path, "rb") as f:
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()

    @staticmethod
    def extract_zip(zip_path: Path, extract_to: Path) -> bool:
        """Extract ZIP file safely, preventing path traversal attacks."""
        try:
            with zipfile.ZipFile(zip_path, "r") as zip_ref:
                for member in zip_ref.namelist():
                    # Check for path traversal
                    member_path = (extract_to / member).resolve()
                    if not str(member_path).startswith(str(extract_to.resolve())):
                        raise ValueError(f"Path traversal attempt detected: {member}")
            zip_ref.extractall(extract_to)
            return True
        except Exception as e:
            raise ValueError(f"Failed to extract ZIP: {e}")
